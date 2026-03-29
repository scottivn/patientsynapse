"""Insurance allowable rates for CPAP/PAP supplies.

Stores payer-specific reimbursement rates by HCPCS code, supply duration,
and effective year. Rates are imported from the annual allowables spreadsheet
and queried during DME order processing for pricing and resupply scheduling.
"""

import logging
from dataclasses import dataclass, field, asdict
from typing import Optional
from datetime import date

import aiosqlite

from server.config import get_settings

logger = logging.getLogger(__name__)


# ── HCPCS code reference for CPAP supplies ────────────────────────

HCPCS_CODES = {
    "A7030": {"description": "Full face mask", "category": "mask"},
    "A7031": {"description": "Full face mask cushion replacement", "category": "mask"},
    "A7032": {"description": "Nasal cushion replacement", "category": "nasal"},
    "A7033": {"description": "Nasal pillow replacement", "category": "nasal"},
    "A7034": {"description": "Nasal mask interface", "category": "nasal"},
    "A7035": {"description": "Headgear", "category": "accessory"},
    "A7036": {"description": "Chinstrap", "category": "accessory"},
    "A7037": {"description": "Tubing", "category": "accessory"},
    "A7038": {"description": "Filter, disposable", "category": "filter"},
    "A7039": {"description": "Filter, non-disposable", "category": "filter"},
    "A7046": {"description": "Water chamber for humidifier", "category": "accessory"},
    "A4604": {"description": "Tubing with heating element", "category": "accessory"},
    "E0601": {"description": "CPAP device", "category": "device"},
    "E0470": {"description": "BiPAP device (without backup rate)", "category": "device"},
    "E0471": {"description": "BiPAP device (with backup rate)", "category": "device"},
}

# Known payer names — used for fuzzy matching during Excel import
KNOWN_PAYERS = [
    "BCBS", "Aetna", "Aetna MC", "Cigna", "UHC", "UHC Medicare",
    "WellMed", "WellMed-UHC", "WellMed-Humana", "Devoted", "Medicare",
    "Humana", "UMR", "Web TPA",
]


@dataclass
class AllowableRate:
    id: int = 0
    payer: str = ""
    payer_plan: str = ""           # "commercial", "medicare_advantage", etc.
    hcpcs_code: str = ""
    description: str = ""
    supply_months: int = 6         # 3 or 6
    allowed_amount: float = 0.0
    effective_year: int = 0
    notes: str = ""
    created_at: str = ""
    updated_at: str = ""


def _get_db_path() -> str:
    settings = get_settings()
    return settings.database_url.replace("sqlite:///", "")


async def init_rates_table():
    """Create allowable_rates table if it doesn't exist."""
    db_path = _get_db_path()
    async with aiosqlite.connect(db_path) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS allowable_rates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                payer TEXT NOT NULL,
                payer_plan TEXT NOT NULL DEFAULT '',
                hcpcs_code TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                supply_months INTEGER NOT NULL DEFAULT 6,
                allowed_amount REAL NOT NULL,
                effective_year INTEGER NOT NULL,
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT NOT NULL DEFAULT (datetime('now')),
                UNIQUE(payer, payer_plan, hcpcs_code, supply_months, effective_year)
            )
        """)
        await db.execute("""
            CREATE INDEX IF NOT EXISTS idx_rates_lookup
            ON allowable_rates(payer, hcpcs_code, effective_year)
        """)
        await db.commit()
    logger.info("Allowable rates table initialized")


# ── Query functions ───────────────────────────────────────────────

async def get_rate(
    payer: str,
    hcpcs_code: str,
    supply_months: int = 6,
    year: Optional[int] = None,
    payer_plan: str = "",
) -> Optional[AllowableRate]:
    """Look up a single allowable rate."""
    if year is None:
        year = date.today().year
    hcpcs_code = _normalize_hcpcs(hcpcs_code)

    db_path = _get_db_path()
    async with aiosqlite.connect(db_path) as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            """SELECT * FROM allowable_rates
               WHERE payer = ? AND hcpcs_code = ? AND supply_months = ?
                 AND effective_year = ? AND payer_plan = ?
               LIMIT 1""",
            (payer, hcpcs_code, supply_months, year, payer_plan),
        )
        row = await cursor.fetchone()
        if not row:
            # Fallback: try without payer_plan
            cursor = await db.execute(
                """SELECT * FROM allowable_rates
                   WHERE payer = ? AND hcpcs_code = ? AND supply_months = ?
                     AND effective_year = ?
                   ORDER BY payer_plan ASC LIMIT 1""",
                (payer, hcpcs_code, supply_months, year),
            )
            row = await cursor.fetchone()
        return _row_to_rate(row) if row else None


async def get_bundle_pricing(
    payer: str,
    hcpcs_codes: list[str],
    supply_months: int = 6,
    year: Optional[int] = None,
    payer_plan: str = "",
) -> dict:
    """Calculate total expected reimbursement for a bundle of HCPCS codes.

    Returns dict with per-item rates and total.
    """
    if year is None:
        year = date.today().year

    items = []
    total = 0.0
    for code in hcpcs_codes:
        rate = await get_rate(payer, code, supply_months, year, payer_plan)
        item = {
            "hcpcs_code": _normalize_hcpcs(code),
            "description": HCPCS_CODES.get(_normalize_hcpcs(code), {}).get("description", ""),
            "allowed_amount": rate.allowed_amount if rate else None,
            "found": rate is not None,
        }
        items.append(item)
        if rate:
            total += rate.allowed_amount

    return {
        "payer": payer,
        "payer_plan": payer_plan,
        "supply_months": supply_months,
        "effective_year": year,
        "items": items,
        "total": round(total, 2),
        "complete": all(i["found"] for i in items),
    }


async def list_rates(
    payer: Optional[str] = None,
    hcpcs_code: Optional[str] = None,
    year: Optional[int] = None,
) -> list[AllowableRate]:
    """List rates with optional filters."""
    db_path = _get_db_path()
    conditions = []
    params = []

    if payer:
        conditions.append("payer = ?")
        params.append(payer)
    if hcpcs_code:
        conditions.append("hcpcs_code = ?")
        params.append(_normalize_hcpcs(hcpcs_code))
    if year:
        conditions.append("effective_year = ?")
        params.append(year)

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    async with aiosqlite.connect(db_path) as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            f"SELECT * FROM allowable_rates {where} ORDER BY payer, hcpcs_code, supply_months",
            params,
        )
        rows = await cursor.fetchall()
        return [_row_to_rate(r) for r in rows]


async def list_payers(year: Optional[int] = None) -> list[dict]:
    """List distinct payers with rate counts."""
    db_path = _get_db_path()
    year_filter = "WHERE effective_year = ?" if year else ""
    params = [year] if year else []

    async with aiosqlite.connect(db_path) as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            f"""SELECT payer, payer_plan, effective_year, COUNT(*) as rate_count
                FROM allowable_rates {year_filter}
                GROUP BY payer, payer_plan, effective_year
                ORDER BY payer, payer_plan""",
            params,
        )
        rows = await cursor.fetchall()
        return [dict(r) for r in rows]


async def get_payer_supply_months(payer: str, hcpcs_code: str, year: Optional[int] = None) -> list[int]:
    """Get available supply month options for a payer + HCPCS code.

    Returns list like [3, 6] — used to determine resupply frequency.
    """
    if year is None:
        year = date.today().year
    hcpcs_code = _normalize_hcpcs(hcpcs_code)

    db_path = _get_db_path()
    async with aiosqlite.connect(db_path) as db:
        cursor = await db.execute(
            """SELECT DISTINCT supply_months FROM allowable_rates
               WHERE payer = ? AND hcpcs_code = ? AND effective_year = ?
               ORDER BY supply_months""",
            (payer, hcpcs_code, year),
        )
        rows = await cursor.fetchall()
        return [r[0] for r in rows]


# ── Write functions ───────────────────────────────────────────────

async def upsert_rate(rate: AllowableRate) -> AllowableRate:
    """Insert or update a single rate."""
    rate.hcpcs_code = _normalize_hcpcs(rate.hcpcs_code)
    if not rate.description:
        rate.description = HCPCS_CODES.get(rate.hcpcs_code, {}).get("description", "")

    db_path = _get_db_path()
    async with aiosqlite.connect(db_path) as db:
        db.row_factory = aiosqlite.Row
        await db.execute(
            """INSERT INTO allowable_rates
               (payer, payer_plan, hcpcs_code, description, supply_months,
                allowed_amount, effective_year, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(payer, payer_plan, hcpcs_code, supply_months, effective_year)
               DO UPDATE SET
                 allowed_amount = excluded.allowed_amount,
                 description = excluded.description,
                 notes = excluded.notes,
                 updated_at = datetime('now')""",
            (rate.payer, rate.payer_plan, rate.hcpcs_code, rate.description,
             rate.supply_months, rate.allowed_amount, rate.effective_year, rate.notes),
        )
        await db.commit()

        # Fetch the inserted/updated row
        cursor = await db.execute(
            """SELECT * FROM allowable_rates
               WHERE payer = ? AND payer_plan = ? AND hcpcs_code = ?
                 AND supply_months = ? AND effective_year = ?""",
            (rate.payer, rate.payer_plan, rate.hcpcs_code,
             rate.supply_months, rate.effective_year),
        )
        row = await cursor.fetchone()
        return _row_to_rate(row) if row else rate


async def bulk_upsert(rates: list[AllowableRate]) -> dict:
    """Insert or update multiple rates in a single transaction."""
    db_path = _get_db_path()
    inserted = 0
    updated = 0

    async with aiosqlite.connect(db_path) as db:
        for rate in rates:
            rate.hcpcs_code = _normalize_hcpcs(rate.hcpcs_code)
            if not rate.description:
                rate.description = HCPCS_CODES.get(rate.hcpcs_code, {}).get("description", "")

            cursor = await db.execute(
                """SELECT id FROM allowable_rates
                   WHERE payer = ? AND payer_plan = ? AND hcpcs_code = ?
                     AND supply_months = ? AND effective_year = ?""",
                (rate.payer, rate.payer_plan, rate.hcpcs_code,
                 rate.supply_months, rate.effective_year),
            )
            existing = await cursor.fetchone()

            await db.execute(
                """INSERT INTO allowable_rates
                   (payer, payer_plan, hcpcs_code, description, supply_months,
                    allowed_amount, effective_year, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(payer, payer_plan, hcpcs_code, supply_months, effective_year)
                   DO UPDATE SET
                     allowed_amount = excluded.allowed_amount,
                     description = excluded.description,
                     notes = excluded.notes,
                     updated_at = datetime('now')""",
                (rate.payer, rate.payer_plan, rate.hcpcs_code, rate.description,
                 rate.supply_months, rate.allowed_amount, rate.effective_year, rate.notes),
            )
            if existing:
                updated += 1
            else:
                inserted += 1

        await db.commit()

    logger.info(f"Bulk upsert: {inserted} inserted, {updated} updated")
    return {"inserted": inserted, "updated": updated, "total": inserted + updated}


async def delete_rate(rate_id: int) -> bool:
    """Delete a rate by ID."""
    db_path = _get_db_path()
    async with aiosqlite.connect(db_path) as db:
        cursor = await db.execute("DELETE FROM allowable_rates WHERE id = ?", (rate_id,))
        await db.commit()
        return cursor.rowcount > 0


async def delete_rates_by_year(year: int) -> int:
    """Delete all rates for a given year (before re-import)."""
    db_path = _get_db_path()
    async with aiosqlite.connect(db_path) as db:
        cursor = await db.execute("DELETE FROM allowable_rates WHERE effective_year = ?", (year,))
        await db.commit()
        count = cursor.rowcount
    logger.info(f"Deleted {count} rates for year {year}")
    return count


# ── Excel import ──────────────────────────────────────────────────

async def import_from_excel(filepath: str, effective_year: Optional[int] = None) -> dict:
    """Parse the insurance allowables Excel file and import rates.

    The spreadsheet is manually maintained with inconsistent layout per payer.
    This parser identifies payer sections, extracts HCPCS codes and rates,
    and flags anything it can't parse for manual review.

    Returns import summary with counts and any parse warnings.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl is required for Excel import. pip install openpyxl"}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    if effective_year is None:
        # Try to extract year from filename
        import re
        match = re.search(r'20\d{2}', filepath)
        effective_year = int(match.group()) if match else date.today().year

    # Read all cells into a grid
    grid = []
    for row in ws.iter_rows(values_only=True):
        grid.append([cell for cell in row])

    if not grid:
        return {"error": "Empty spreadsheet", "rates": []}

    rates = []
    warnings = []

    # Strategy: scan for payer section headers, then extract rates below
    # A payer header cell contains a known payer name + supply info (e.g., "BCBS", "Cigna 6 month ffm")
    sections = _identify_sections(grid, warnings)

    for section in sections:
        section_rates = _extract_section_rates(grid, section, effective_year, warnings)
        rates.extend(section_rates)

    # Deduplicate — keep highest rate if duplicates
    seen = {}
    deduped = []
    for r in rates:
        key = (r.payer, r.payer_plan, r.hcpcs_code, r.supply_months)
        if key in seen:
            if r.allowed_amount > seen[key].allowed_amount:
                seen[key] = r
        else:
            seen[key] = r
            deduped.append(r)

    # Update deduped list with best values
    final_rates = list(seen.values())

    # Persist to database
    if final_rates:
        result = await bulk_upsert(final_rates)
    else:
        result = {"inserted": 0, "updated": 0, "total": 0}

    return {
        "effective_year": effective_year,
        "rates_parsed": len(final_rates),
        "payer_sections_found": len(sections),
        "payers": list(set(r.payer for r in final_rates)),
        "hcpcs_codes": sorted(set(r.hcpcs_code for r in final_rates)),
        "warnings": warnings[:50],  # Cap warnings
        **result,
    }


# ── Parser internals ─────────────────────────────────────────────

import re

# Matches "A7030", "E0601", etc. (letter + 4 digits)
_HCPCS_PREFIXED = re.compile(r'([A-Za-z])(\d{4})')
# Matches bare 4-digit codes common in this spreadsheet: "7030", "4604"
_HCPCS_BARE = re.compile(r'^(\d{4})$')
_AMOUNT_PATTERN = re.compile(r'(\d+\.?\d*)')

# Payer name patterns for header detection
_PAYER_PATTERNS = [
    (re.compile(r'BCBS', re.I), "BCBS", ""),
    (re.compile(r'Aetna\s*MC', re.I), "Aetna", "medicare_advantage"),
    (re.compile(r'Aetna\s*reg', re.I), "Aetna", "commercial"),
    (re.compile(r'Aetna', re.I), "Aetna", "commercial"),
    (re.compile(r'Cigna', re.I), "Cigna", ""),
    (re.compile(r'UHC\s*Medicare\s*AD', re.I), "UHC", "medicare_advantage"),
    (re.compile(r'UHC\s*Med', re.I), "UHC", "medicare_advantage"),
    (re.compile(r'MEDICARE\s*\d\s*mon', re.I), "Medicare", ""),
    (re.compile(r'MEDICARE', re.I), "Medicare", ""),
    (re.compile(r'MED\s+\d\s*MON', re.I), "Medicare", ""),
    (re.compile(r'WellMed[\s-]*UHC', re.I), "WellMed-UHC", ""),
    (re.compile(r'WellMed[\s-]*Humana', re.I), "WellMed-Humana", ""),
    (re.compile(r'WellMed', re.I), "WellMed", ""),
    (re.compile(r'Devoted', re.I), "Devoted", ""),
    (re.compile(r'Humana', re.I), "Humana", ""),
    (re.compile(r'UMR', re.I), "UMR", ""),
    (re.compile(r'Web\s*TPA', re.I), "Web TPA", ""),
    (re.compile(r'UHC\s*ffm', re.I), "UHC", "commercial"),
    (re.compile(r'UHC\s*nasal', re.I), "UHC", "commercial"),
    (re.compile(r'uhc\b', re.I), "UHC", ""),
]

# Equipment type patterns for determining mask type context
_EQUIP_PATTERNS = [
    (re.compile(r'ffm|full\s*face', re.I), "ffm"),
    (re.compile(r'nasal\s*cushion|cushion', re.I), "nasal_cushion"),
    (re.compile(r'nasal\s*pillow|pillow', re.I), "nasal_pillow"),
    (re.compile(r'nasal', re.I), "nasal"),
]

_SUPPLY_PATTERN = re.compile(r'(\d)\s*(?:mon|m\b|month)', re.I)

# Sub-header patterns: cells that describe supply/type but no payer name
_SUBHEADER_PATTERN = re.compile(
    r'(?:\d\s*(?:mon|m\b|month)|ffm|full\s*face|nasal|pillow|cushion)',
    re.I,
)


def _parse_hcpcs(cell_str: str) -> str | None:
    """Extract an HCPCS code from a cell string. Handles 'A7030', 'a7030', or bare '7030'."""
    m = _HCPCS_PREFIXED.search(cell_str)
    if m:
        return f"{m.group(1).upper()}{m.group(2)}"
    # Bare 4-digit code in its own cell (column 0 / key column pattern)
    bare = cell_str.strip()
    if _HCPCS_BARE.match(bare):
        return f"A{bare}"
    return None


def _identify_sections(grid: list, warnings: list) -> list[dict]:
    """Scan the grid for payer section headers.

    Handles two layouts:
    1. Payer name + supply/type in one cell (e.g., "Aetna reg 6 month ffm")
    2. Payer name alone (e.g., "BCBS") with sub-headers in adjacent cells
       (e.g., "6 month ffm", "3 month nasal"). These inherit the payer.
    """
    sections = []
    if not grid:
        return sections

    for row_idx, row in enumerate(grid):
        # Track the last payer found on this row for sub-header inheritance
        row_payer = None
        row_plan = ""
        row_payer_col = -1

        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if len(cell_str) < 2 or len(cell_str) > 80:
                continue

            # Check if cell matches a known payer
            matched_payer = False
            for pattern, payer, plan in _PAYER_PATTERNS:
                if pattern.search(cell_str):
                    supply_match = _SUPPLY_PATTERN.search(cell_str)
                    supply_months = int(supply_match.group(1)) if supply_match else 6

                    equip_type = ""
                    for ep, etype in _EQUIP_PATTERNS:
                        if ep.search(cell_str):
                            equip_type = etype
                            break

                    # Only add as section if it has supply/type info OR is a payer-only header
                    has_detail = bool(supply_match) or bool(equip_type)
                    if has_detail:
                        sections.append({
                            "payer": payer, "payer_plan": plan,
                            "col": col_idx, "header_row": row_idx,
                            "supply_months": supply_months, "equip_type": equip_type,
                            "header_text": cell_str[:60],
                        })
                    # Track as row payer for sub-header inheritance
                    row_payer = payer
                    row_plan = plan
                    row_payer_col = col_idx
                    matched_payer = True
                    break

            # If no payer matched, check if this is a sub-header (supply/type only)
            # that should inherit the payer from the left
            if not matched_payer and row_payer and col_idx > row_payer_col:
                if _SUBHEADER_PATTERN.search(cell_str):
                    supply_match = _SUPPLY_PATTERN.search(cell_str)
                    supply_months = int(supply_match.group(1)) if supply_match else 6

                    equip_type = ""
                    for ep, etype in _EQUIP_PATTERNS:
                        if ep.search(cell_str):
                            equip_type = etype
                            break

                    sections.append({
                        "payer": row_payer, "payer_plan": row_plan,
                        "col": col_idx, "header_row": row_idx,
                        "supply_months": supply_months, "equip_type": equip_type,
                        "header_text": f"{row_payer}: {cell_str[:50]}",
                    })

    logger.info(f"Excel import: found {len(sections)} payer section headers")
    return sections


def _find_hcpcs_columns(grid: list, header_row: int, scan_rows: int = 15) -> list[int]:
    """Find columns that contain HCPCS codes (key columns) near a header row."""
    hcpcs_cols = set()
    for row_idx in range(header_row + 1, min(header_row + scan_rows, len(grid))):
        row = grid[row_idx]
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            code = _parse_hcpcs(str(cell).strip())
            if code and code[0] == "A":
                # If the cell is JUST a code (no amount mixed in), it's a key column
                cell_str = str(cell).strip()
                if _HCPCS_BARE.match(cell_str) or len(cell_str) <= 5:
                    hcpcs_cols.add(col_idx)
    return sorted(hcpcs_cols)


def _extract_section_rates(
    grid: list,
    section: dict,
    effective_year: int,
    warnings: list,
) -> list[AllowableRate]:
    """Extract rates from a single payer section column."""
    rates = []
    col = section["col"]
    start_row = section["header_row"] + 1
    payer = section["payer"]
    payer_plan = section["payer_plan"]
    supply_months = section["supply_months"]

    # Find HCPCS key columns near this section
    hcpcs_key_cols = _find_hcpcs_columns(grid, section["header_row"])

    for row_idx in range(start_row, min(start_row + 15, len(grid))):
        row = grid[row_idx]
        if col >= len(row):
            continue

        cell_val = row[col]
        if cell_val is None:
            continue
        cell_str = str(cell_val).strip()
        if not cell_str:
            continue

        hcpcs = None
        amount = None

        # Check if cell has a code + amount combined (e.g., "A7034 63.25" or "a7030 56")
        hcpcs = _parse_hcpcs(cell_str)
        if hcpcs:
            # Extract amount from same cell if present
            amounts = _AMOUNT_PATTERN.findall(cell_str)
            # Filter out the HCPCS digits from the amounts
            hcpcs_digits = hcpcs[1:]  # e.g., "7030"
            for a in reversed(amounts):
                if a != hcpcs_digits and a not in hcpcs_digits:
                    try:
                        val = float(a)
                        if 0 < val < 5000:
                            amount = val
                            break
                    except ValueError:
                        pass

        # Check if cell is a pure number (rate value)
        if hcpcs is None:
            try:
                val = float(cell_str.replace(",", ""))
                if 0 < val < 5000:
                    amount = val
            except (ValueError, TypeError):
                amounts = _AMOUNT_PATTERN.findall(cell_str)
                if amounts:
                    try:
                        val = float(amounts[-1])
                        if 0 < val < 5000:
                            amount = val
                    except ValueError:
                        pass

        # If we have a code but no amount, check the adjacent column to the right
        if hcpcs and amount is None and col + 1 < len(row):
            adj = row[col + 1]
            if adj is not None:
                try:
                    val = float(str(adj).strip().replace(",", ""))
                    if 0 < val < 5000:
                        amount = val
                except (ValueError, TypeError):
                    pass

        # If we have an amount but no code, look in HCPCS key columns for this row
        if amount is not None and hcpcs is None:
            for key_col in hcpcs_key_cols:
                if key_col < len(row) and row[key_col] is not None:
                    code = _parse_hcpcs(str(row[key_col]).strip())
                    if code:
                        hcpcs = code
                        break
            # Fallback: check columns to the left (up to 3)
            if hcpcs is None:
                for check_col in range(max(0, col - 3), col):
                    if check_col < len(row) and row[check_col] is not None:
                        code = _parse_hcpcs(str(row[check_col]).strip())
                        if code:
                            hcpcs = code
                            break

        # If we found both, create a rate
        if hcpcs and amount is not None and amount > 0:
            rate = AllowableRate(
                payer=payer,
                payer_plan=payer_plan,
                hcpcs_code=hcpcs,
                supply_months=supply_months,
                allowed_amount=round(amount, 2),
                effective_year=effective_year,
            )
            rates.append(rate)

    if not rates:
        warnings.append(f"No rates extracted for {payer} ({section.get('header_text', '')})")

    return rates


def _normalize_hcpcs(code: str) -> str:
    """Normalize HCPCS code to uppercase. Add A prefix only for bare numeric codes."""
    code = code.strip().upper()
    if code and code[0].isdigit():
        code = f"A{code}"
    return code


def _row_to_rate(row) -> AllowableRate:
    """Convert a database row to an AllowableRate dataclass."""
    return AllowableRate(
        id=row["id"],
        payer=row["payer"],
        payer_plan=row["payer_plan"],
        hcpcs_code=row["hcpcs_code"],
        description=row["description"],
        supply_months=row["supply_months"],
        allowed_amount=row["allowed_amount"],
        effective_year=row["effective_year"],
        notes=row["notes"],
        created_at=row["created_at"],
        updated_at=row["updated_at"],
    )
